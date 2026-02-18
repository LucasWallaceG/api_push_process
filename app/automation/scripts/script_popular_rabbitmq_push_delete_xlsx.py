import os
import time
import json
import requests
import pandas as pd
from datetime import date
from  Models.scripts_pje import *
from openpyxl import load_workbook
from script_acessar_pje import preparar_ambiente
from pagepush.script_data_push import AutomacaoPush, ProcessosScraper


URL_AUTOMACAO = "http://192.168.11.38:9000/automacao/rabbit/push/delet/"
DELAY = 5
SHEET = r"Tag_evernote_excluir_push/Evernote.xlsx"
PLANILHA_EVERNOTE = "Tag_evernote_excluir_push/Evernote_30012026.xlsx"

df = pd.read_excel(PLANILHA_EVERNOTE)


def extrair_trt_do_processo(numero_processo):
    partes = numero_processo.split(".")
    return str(int(partes[3]))


def obter_pagina_do_json(trt, processo):
    hoje = date.today().isoformat()
    # path = f"processos_push_{hoje}.json"

    # manual:
    path = "others/processos_push_2026-01-30_v1.json"

    if not os.path.exists(path):
        return None

    with open(path, "r", encoding="utf-8") as f:
        dados = json.load(f)

    for item in dados:
        if (
            item.get("tribunal") == f"TRT{trt}"
            and item.get("numero_processo") == processo
        ):
            return str(item.get("pagina"))

    return None


def atualizar_planilha(processo, status, mensagem):
    wb = load_workbook(PLANILHA_EVERNOTE)
    ws = wb.active  # ou ws = wb["Sheet1"]

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip() == processo:
            row[2].value = status      # STATUS
            row[3].value = mensagem    # MENSAGEM
            break

    wb.save(PLANILHA_EVERNOTE)


def normalizar_resultado(retorno):
    """
    Normaliza o retorno da automação.
    Aceita: dict | str | None
    Retorna: (RESULTADO, mensagem_normalizada)
    """

    # 🔹 Caso None
    if retorno is None:
        return "ERRO", "Erro não detalhado"

    # 🔹 Caso dict (retorno padrão da automação)
    if isinstance(retorno, dict):
        sucesso = bool(retorno.get("sucesso"))
        texto = retorno.get("mensagem") or ""

    # 🔹 Caso string ou outro tipo
    else:
        texto = str(retorno)
        sucesso = "sucesso" in texto.lower()

    texto = texto.strip()
    texto_lower = texto.lower()

    # 🔥 REGRA DE NEGÓCIO
    if "já cadastrado" in texto_lower or "ja cadastrado" in texto_lower:
        return "AVISO", "Processo já cadastrado"

    if sucesso:
        return "SUCESSO", texto or "Operação realizada com sucesso"

    if "erro" in texto_lower or "falha" in texto_lower or "não encontrado" in texto_lower:
        return "ERRO", texto or "Erro retornado pela automação"

    # fallback conservador
    return "ERRO", texto or "Erro não detalhado"


print("🧠 Criando driver Selenium...")
driver = criar_driver()
automation = AutomacaoPush(driver)

for _, row in df.iterrows():
    numero_processo = str(row["NUMERO_PROCESSO"]).strip()

    status = str(row.get("STATUS", "")).strip().lower()

    # ⏭ pula se já foi processado com sucesso
    if status == "sucesso" or status == "não encontrado":
        print(f"⏭ Já processado com sucesso: {numero_processo}")
        continue

    try:
        trt = str(int(numero_processo.split(".")[3]))
    except Exception:
        print(f"❌ TRT inválido: {numero_processo}")
        continue


    pagina = obter_pagina_do_json(trt, numero_processo)
    # trt = extrair_trt_do_processo(numero_processo)

    if not pagina:
        print(f"⚠️ Página não encontrada no JSON: {numero_processo}")

        atualizar_planilha(
            numero_processo,
            "FALHA",
            "Página não encontrada no JSON"
        )

        continue

    print(f'- [STATUS]: {status}')
    print(f'- [TRT]: {trt}')
    print(f'- [PÁG.]: {pagina}')
    print(f'- [Nº PROCESSO]: {numero_processo}')

    context = preparar_ambiente(
        driver=driver,
        automation=automation,
        trt=trt,
        perito="paula"
    )

    mensagem = automation.function_main_del_push(pagina, numero_processo, context)
    resultado, mensagem_normalizada = normalizar_resultado(mensagem)
    print(f'- [RETURN]: {resultado} | {mensagem_normalizada}')

    if resultado in ("SUCESSO", "AVISO"):
        atualizar_planilha(numero_processo, "SUCESSO", mensagem_normalizada)
    else:
        atualizar_planilha(numero_processo, "NÃO ENCONTRADO", mensagem_normalizada)
