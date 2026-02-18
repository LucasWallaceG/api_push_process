import os
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook


ARQUIVO_LOG = "log_push.xlsx"


def salvar_log_push_xlsx(
    tribunal,
    numero_processo,
    resultado,
    mensagem,
):
    data_extracao = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if os.path.exists(ARQUIVO_LOG):
        wb = load_workbook(ARQUIVO_LOG)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "data_extracao",
            "tribunal",
            "numero_processo",
            "resultado",
            "mensagem"
        ])

    ws.append([
        data_extracao,
        tribunal,
        numero_processo,
        resultado,
        mensagem
    ])

    wb.save(ARQUIVO_LOG)
    wb.close()  # 🔥 sempre fechar


def salvar_log_push_xlsx_v0(tribunal, numero_processo, mensagem):
    data_extracao = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    status = (
        "Sucesso"
        if mensagem and ("sucesso" or "Processo já cadastrado") in mensagem.lower()
        else "Erro"
    )

    if os.path.exists(ARQUIVO_LOG):
        wb = load_workbook(ARQUIVO_LOG)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "data_extracao",
            "tribunal",
            "numero_processo",
            "mensagem",
            "status"
        ])

    ws.append([
        data_extracao,
        tribunal,
        numero_processo,
        mensagem,
        status
    ])

    wb.save(ARQUIVO_LOG)
    wb.close()  # 🔥 FUNDAMENTAL


def salvar_log_trts_csv(log_trts, path="log_trts_push.csv"):
    existe = os.path.exists(path)

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        if not existe:
            writer.writerow(["data_execucao", "tribunal", "quantidade"])

        data_execucao = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for tribunal, qtd in log_trts.items():
            writer.writerow([data_execucao, tribunal, qtd])
