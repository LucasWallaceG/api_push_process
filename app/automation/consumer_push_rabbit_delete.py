import os
import json
import pika
from datetime import date
from  Models.scripts_pje import *
from openpyxl import load_workbook
from script_acessar_pje import preparar_ambiente
from pagepush.script_data_push import AutomacaoPush, ProcessosScraper
from pagepush.script_data_push import AutomacaoPush, ProcessosScraper
from script_salvar_log_return_csv import salvar_log_push_xlsx as save_log


# ================== CONFIG ==================
RABBIT_HOST = "192.168.11.38"
QUEUE_NAME = "q.push_delet"
RETRY_DELAY = 5
first = True

RABBIT_USER = "ro.berto"
RABBIT_PASS = "Jrs-2018"

PLANILHA_EVERNOTE = "Tag_evernote_excluir_push/Evernote.xlsx"

# 🔥 TRTs que NÃO devem ser processados
# TRT_EXCECOES = {'5'}   # set é mais rápido que list
TRT_EXCECOES = {}   # set é mais rápido que list

# ===========================================

TRT_COOLDOWN = {}  # { trt: datetime }
TRT_COOLDOWN_MINUTES = 10


def atualizar_planilha(processo, status, mensagem):
    wb = load_workbook(PLANILHA_EVERNOTE)
    ws = wb.active  # ou ws = wb["Sheet1"]

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip() == processo:
            row[2].value = status      # STATUS
            row[3].value = mensagem    # MENSAGEM
            break

    wb.save(PLANILHA_EVERNOTE)


def localizar_e_deletar_processo(automation, trt, processo):
    """
    Usa JSON diário como atalho.
    Retorna (True, mensagem) ou (False, mensagem)
    """

    json_paginas = carregar_json_paginacao(trt)
    pagina_sugerida = json_paginas.get(processo)

    scraper = ProcessosScraper(
        automation.driver,
        trt,
        chaves_unicas=set(),
        data_execucao=None
    )

    # 🔹 tenta via JSON
    if pagina_sugerida:
        print(f"[INFO] Página sugerida pelo JSON: {pagina_sugerida}")
        automation.ir_para_pagina_json(pagina_sugerida)

        achado = scraper.localizar_processo(processo)
        if achado:
            automation.deletar_linha(achado["row"])
            return True, "Deletado com base na página do JSON"

    # 🔹 fallback: busca sequencial
    print("[INFO] Fallback: busca completa na tabela")
    scraper.page = 1
    achado = scraper.localizar_processo(processo)

    if achado:
        automation.deletar_linha(achado["row"])
        return True, "Deletado após busca sequencial"

    return False, "Processo não localizado na tabela"


def carregar_json_paginacao(trt):
    """
    Carrega o JSON diário e retorna um dict:
    { numero_processo: pagina }
    """
    hoje = date.today().isoformat()
    path = f"processos_push_{hoje}.json"

    if not os.path.exists(path):
        return {}

    with open(path, "r", encoding="utf-8") as f:
        dados = json.load(f)

    mapa = {}
    for item in dados:
        if item.get("tribunal") == f"TRT{trt}":
            mapa[item["numero_processo"]] = int(item["pagina"])

    return mapa


def extrair_trt_do_processo(numero_processo):
    partes = numero_processo.split(".")
    return str(int(partes[3]))


def normalizar_resultado(retorno):
    """
    Normaliza o retorno da automação.
    Aceita: dict | str | None
    Retorna: (RESULTADO, mensagem_normalizada)
    """

    # 🔹 Caso None
    if retorno is None:
        return "ERRO", "Erro não detalhado"

    # 🔹 Caso dict (retorno padrão da automação)
    if isinstance(retorno, dict):
        sucesso = bool(retorno.get("sucesso"))
        texto = retorno.get("mensagem") or ""

    # 🔹 Caso string ou outro tipo
    else:
        texto = str(retorno)
        sucesso = "sucesso" in texto.lower()

    texto = texto.strip()
    texto_lower = texto.lower()

    # 🔥 REGRA DE NEGÓCIO
    if "já cadastrado" in texto_lower or "ja cadastrado" in texto_lower:
        return "AVISO", "Processo já cadastrado"

    if sucesso:
        return "SUCESSO", texto or "Operação realizada com sucesso"

    if "erro" in texto_lower or "falha" in texto_lower or "não encontrado" in texto_lower:
        return "ERRO", texto or "Erro retornado pela automação"

    # fallback conservador
    return "ERRO", texto or "Erro não detalhado"


def trt_em_cooldown(trt):
    ate = TRT_COOLDOWN.get(trt)
    if not ate:
        return False
    return datetime.now() < ate


def callback(ch, method, properties, body):
    global automation, driver, first

    processo = None
    trt = None

    try:
        data = json.loads(body.decode())
        processo = data.get("numero_processo")

        print(f"\n📥 Processo recebido da fila: {processo}")

        trt = extrair_trt_do_processo(processo)

        # 🚫 VERIFICA EXCEÇÕES
        if trt in TRT_EXCECOES:
            print(f"⏭ TRT {trt} está na lista de exceções. Processo {processo} ignorado.")
            ch.basic_ack(delivery_tag=method.delivery_tag)  # 🔥 REMOVE DA FILA
            return

        context = preparar_ambiente(
            driver=driver,
            automation=automation,
            trt=trt,
            perito="paula"
        )

        mensagem = automation.function_main_del_push(processo, context)
        resultado, mensagem_normalizada = normalizar_resultado(mensagem)
        print(f'- (Result): {resultado} | {mensagem_normalizada}')

        if resultado in ("SUCESSO" or "AVISO"):
            atualizar_planilha(processo, "SUCESSO", mensagem)
        else:
            atualizar_planilha(processo, "NÃO ENCONTRADO", mensagem)

        # save_log(
        #     tribunal=f"TRT{trt}",
        #     numero_processo=processo,
        #     resultado=resultado,
        #     mensagem=mensagem_normalizada,
        # )

        if resultado in ("SUCESSO" or "AVISO"):
            # ✅ remove da fila (inclusive se já estava cadastrado)
            ch.basic_ack(delivery_tag=method.delivery_tag)
        else:
            # ❌ erro real → volta para fila
            ch.basic_nack(delivery_tag=method.delivery_tag, requeue=True)

    except Exception as e:

        erro_txt = f"Exceção não tratada: {e}"
        print(f"❌ {erro_txt}")

        # 🔥 LOGA O ERRO CRÍTICO
        save_log(
            tribunal=trt,
            numero_processo=processo,
            resultado="ERRO",
            mensagem=erro_txt,
        )

        # ❌ erro real → volta para fila
        ch.basic_nack(
            delivery_tag=method.delivery_tag,
            requeue=True
        )

        time.sleep(RETRY_DELAY)


def start_consumer():
    global automation
    global driver

    print("🧠 Criando driver Selenium...")
    driver = criar_driver()

    automation = AutomacaoPush(driver)

    credentials = pika.PlainCredentials(
        username=RABBIT_USER,
        password=RABBIT_PASS
    )

    params = pika.ConnectionParameters(
        host=RABBIT_HOST,
        port=5672,
        virtual_host="/myvhost",  # ✅ VHOST CORRETO
        credentials=credentials,
        heartbeat=600,
        blocked_connection_timeout=300
    )

    connection = pika.BlockingConnection(params)
    channel = connection.channel()

    channel.queue_declare(queue=QUEUE_NAME, durable=True)
    channel.basic_qos(prefetch_count=1)

    channel.basic_consume(
        queue=QUEUE_NAME,
        on_message_callback=callback
    )

    print("🐇 Aguardando mensagens na fila...")
    channel.start_consuming()



if __name__ == "__main__":
    start_consumer()
